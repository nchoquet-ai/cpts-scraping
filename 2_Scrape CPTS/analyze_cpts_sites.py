"""
Analyse comparative des sites CPTS
====================================
Visite chaque site, détecte la structure des données et génère un rapport Excel.

Usage :
  pip install httpx pandas openpyxl tqdm
  python analyze_cpts_sites.py --input "Test_sites_internet_cpts.xlsx"
  python analyze_cpts_sites.py --input "..\\2_Scrape CPTS\\Input\\Test.xlsx" --output "..\\2_Scrape CPTS\\Output"

Le rapport Excel généré (Output/analyse_sites.xlsx) indique pour chaque site :
  - Accessibilité HTTP
  - Présence de tableaux HTML (équipe, adhérents...)
  - Nombre et URLs de PDFs liés
  - Nombre d'images (trombinoscopes potentiels)
  - Pages "équipe/bureau/gouvernance" détectées dans la navigation
  - Mots-clés CPTS trouvés dans le texte
  - Recommandation de stratégie d'extraction
"""

import argparse
import asyncio
import re
import time
from pathlib import Path
from urllib.parse import urljoin, urlparse

import httpx
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from tqdm import tqdm

OUTPUT_DIR = Path("Output")  # remplacé dynamiquement via --output

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
}

# Mots-clés pour détecter les pages équipe dans la navigation
EQUIPE_KW = [
    "equipe", "bureau", "gouvernance", "membre", "president", "conseil",
    "administration", "qui-sommes", "qui_sommes", "organisation", "instance",
    "direction", "ca-", "comite", "dirigeant", "contact",
]

# Mots-clés CPTS à chercher dans le texte des pages
CONTENT_KW = [
    "président", "vice-président", "trésorier", "secrétaire", "bureau",
    "adhérent", "adhérents", "médecin", "infirmier", "infirmière",
    "kinésithérapeute", "pharmacien", "sage-femme", "projet de santé",
    "mission", "territoire", "communes",
]

# Mots-clés pour détecter les actus
ACTU_KW = ["actualit", "news", "agenda", "evenement", "événement", "blog", "article"]


def normalize_url(raw: str) -> str:
    s = str(raw).strip().rstrip("/")
    if not s or s.lower() in {"nan", "none", "", "n/a"}:
        return ""
    if s.startswith(("http://", "https://")):
        return s
    return "https://" + s.lstrip("www.")


def extract_text(html: str) -> str:
    html = re.sub(r"<script[^>]*>.*?</script>", " ", html, flags=re.DOTALL | re.I)
    html = re.sub(r"<style[^>]*>.*?</style>", " ", html, flags=re.DOTALL | re.I)
    text = re.sub(r"<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", text).strip()


def find_internal_links(html: str, base_url: str):
    """Retourne (links_internes, pdfs, liens_images)."""
    base_domain = urlparse(base_url).netloc
    found = re.findall(r'href=["\']([^"\'#\s]+)["\']', html, re.I)
    internal, pdfs = [], []
    for href in found:
        abs_url = urljoin(base_url, href)
        parsed = urlparse(abs_url)
        if parsed.netloc and parsed.netloc != base_domain:
            continue
        if abs_url.lower().endswith(".pdf"):
            pdfs.append(abs_url)
        else:
            internal.append(abs_url)
    return list(dict.fromkeys(internal))[:50], list(dict.fromkeys(pdfs))[:20]


def score_link(url: str) -> int:
    url_lower = url.lower()
    return sum(1 for k in EQUIPE_KW if k in url_lower)


def detect_actu_links(links: list) -> list:
    return [l for l in links if any(k in l.lower() for k in ACTU_KW)]


def count_img_tags(html: str) -> int:
    return len(re.findall(r"<img\b", html, re.I))


def detect_strategy(result: dict) -> str:
    """Déduit la stratégie d'extraction recommandée."""
    if result.get("http_status") == 403:
        return "⚠️ Accès bloqué (Playwright requis)"
    if (result.get("http_status") or 0) >= 400:
        return "❌ Site inaccessible"

    strategies = []

    if result["has_table"]:
        strategies.append("📊 Tableau HTML → extraction directe")
    if result["pdfs"]:
        strategies.append(f"📄 PDF ({len(result['pdfs'])}) → pdfplumber + Claude")
    if result["img_count"] > 5 and not result["has_table"]:
        strategies.append("🖼️ Images → Claude Vision (coûteux)")
    if result["equipe_pages"]:
        strategies.append(f"🔗 Page dédiée équipe ({len(result['equipe_pages'])} détectée(s))")
    if result["actu_pages"]:
        strategies.append(f"📰 Page actus ({len(result['actu_pages'])} détectée(s))")
    if result["kw_found"]:
        strategies.append("✅ Texte extractible directement")

    return " | ".join(strategies) if strategies else "🔍 Analyse manuelle requise"


async def analyze_site(client: httpx.AsyncClient, code: str, nom: str, url: str) -> dict:
    result = {
        "code": code,
        "nom": nom,
        "url_input": url,
        "http_status": None,
        "url_finale": "",
        "has_table": False,
        "pdfs": [],
        "img_count": 0,
        "equipe_pages": [],
        "actu_pages": [],
        "kw_found": [],
        "nb_liens_internes": 0,
        "texte_preview": "",
        "erreur": "",
        "strategie": "",
    }

    if not url:
        result["erreur"] = "URL vide"
        result["strategie"] = "❌ Pas d'URL"
        return result

    try:
        r = await client.get(url, follow_redirects=True, timeout=15)
        result["http_status"] = r.status_code
        result["url_finale"] = str(r.url)

        if r.status_code == 403:
            result["strategie"] = "⚠️ Accès bloqué (Playwright requis)"
            return result

        if r.status_code >= 400:
            result["strategie"] = f"❌ HTTP {r.status_code}"
            return result

        html = r.text
        text = extract_text(html)
        result["texte_preview"] = text[:600]

        # Structure
        result["has_table"] = bool(re.search(r"<table\b", html, re.I))
        result["img_count"] = count_img_tags(html)

        # Liens
        internal_links, pdfs = find_internal_links(html, str(r.url))
        result["nb_liens_internes"] = len(internal_links)
        result["pdfs"] = pdfs

        # Pages équipe
        equipe = [(l, score_link(l)) for l in internal_links if score_link(l) > 0]
        equipe.sort(key=lambda x: -x[1])
        result["equipe_pages"] = [l for l, _ in equipe[:5]]

        # Pages actus
        result["actu_pages"] = detect_actu_links(internal_links)[:3]

        # Mots-clés trouvés
        result["kw_found"] = [k for k in CONTENT_KW if k.lower() in text.lower()]

    except httpx.ConnectTimeout:
        result["erreur"] = "Timeout connexion"
    except httpx.ReadTimeout:
        result["erreur"] = "Timeout lecture"
    except Exception as e:
        result["erreur"] = str(e)[:120]

    result["strategie"] = detect_strategy(result)
    return result


async def run(input_path: str, output_dir: Path = None):
    global OUTPUT_DIR
    if output_dir:
        OUTPUT_DIR = output_dir
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    df_input = pd.read_excel(input_path, dtype=str).fillna("")

    # Normaliser les colonnes
    col_map = {c.strip(): c.strip() for c in df_input.columns}
    df_input.columns = [c.strip() for c in df_input.columns]
    url_col = next((c for c in df_input.columns if "site" in c.lower()), df_input.columns[2])
    nom_col = next((c for c in df_input.columns if "nom" in c.lower()), df_input.columns[1])
    code_col = df_input.columns[0]

    sites = []
    for _, row in df_input.iterrows():
        url = normalize_url(row.get(url_col, ""))
        sites.append((row[code_col], row[nom_col], url))

    print(f"\n{len(sites)} sites à analyser...\n")

    results = []
    async with httpx.AsyncClient(headers=HEADERS, follow_redirects=True) as client:
        tasks = [analyze_site(client, c, n, u) for c, n, u in sites]
        for coro in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc="Analyse"):
            results.append(await coro)

    # Trier dans l'ordre original
    order = {c: i for i, (c, _, _) in enumerate(sites)}
    results.sort(key=lambda r: order.get(r["code"], 999))

    export_excel(results, df_input)
    print(f"\n✅ Rapport généré : {(OUTPUT_DIR / 'analyse_sites.xlsx').resolve()}")
    print_summary(results)


def export_excel(results: list, df_input: pd.DataFrame):
    out = OUTPUT_DIR / "analyse_sites.xlsx"

    rows = []
    for r in results:
        rows.append({
            "Code CPTS":           r["code"],
            "Nom CPTS":            r["nom"],
            "URL":                 r["url_input"],
            "HTTP Status":         r["http_status"] or r["erreur"],
            "URL finale":          r["url_finale"],
            "Tableau HTML":        "OUI" if r["has_table"] else "non",
            "Nb PDFs liés":        len(r["pdfs"]),
            "URLs PDFs":           "\n".join(r["pdfs"][:5]),
            "Nb images":           r["img_count"],
            "Pages équipe/bureau": "\n".join(r["equipe_pages"]),
            "Pages actus":         "\n".join(r["actu_pages"]),
            "Mots-clés trouvés":   ", ".join(r["kw_found"]),
            "Stratégie recommandée": r["strategie"],
            "Texte extrait (preview)": r["texte_preview"][:400],
            "Erreur":              r["erreur"],
        })

    df = pd.DataFrame(rows)
    df.to_excel(out, index=False)

    wb = load_workbook(out)
    ws = wb.active

    # En-tête
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Largeurs
    widths = [12, 35, 42, 10, 42, 12, 10, 55, 10, 55, 45, 40, 55, 60, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # Couleurs par statut
    COLOR_OK      = "C6EFCE"
    COLOR_BLOCKED = "FCE4D6"
    COLOR_ERROR   = "FFC7CE"
    COLOR_PARTIAL = "FFEB9C"

    status_col = 4  # HTTP Status (col D)
    strat_col  = 13  # Stratégie (col M)

    for row in ws.iter_rows(min_row=2):
        status = str(row[status_col - 1].value or "")
        strat  = str(row[strat_col - 1].value or "")

        if "200" in status or "301" in status or "302" in status:
            color = COLOR_OK
        elif "403" in status or "Bloqué" in strat:
            color = COLOR_BLOCKED
        elif any(x in status for x in ["404", "500", "Timeout", "Error"]):
            color = COLOR_ERROR
        else:
            color = COLOR_PARTIAL

        fill = PatternFill("solid", fgColor=color)
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Hauteur des lignes
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 80

    wb.save(out)


def print_summary(results: list):
    total = len(results)
    ok    = sum(1 for r in results if r["http_status"] and 200 <= r["http_status"] < 400)
    blocked = sum(1 for r in results if r["http_status"] == 403)
    error   = sum(1 for r in results if r.get("erreur") and not r["http_status"])
    tables  = sum(1 for r in results if r["has_table"])
    pdfs    = sum(1 for r in results if r["pdfs"])
    imgs    = sum(1 for r in results if r["img_count"] > 5 and not r["has_table"])
    equipe  = sum(1 for r in results if r["equipe_pages"])

    print("\n─── Résumé ─────────────────────────────────────────────")
    print(f"  Total sites analysés       : {total}")
    print(f"  ✅ Accessibles             : {ok}")
    print(f"  🔒 Bloqués (403)           : {blocked}")
    print(f"  ❌ Inaccessibles/erreur    : {error}")
    print(f"  ─── Structure ───────────────────────────────────────")
    print(f"  📊 Tableau HTML détecté    : {tables}")
    print(f"  📄 PDFs liés               : {pdfs}")
    print(f"  🖼️  Images (sans tableau)   : {imgs}")
    print(f"  🔗 Page équipe détectée    : {equipe}")
    print("────────────────────────────────────────────────────────\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Analyse des sites CPTS")
    parser.add_argument("--input",  required=True, help="Fichier Excel source")
    parser.add_argument("--output", default=None,  help="Dossier de sortie (défaut: Output/)")
    args = parser.parse_args()
    output_dir = Path(args.output) if args.output else None
    asyncio.run(run(args.input, output_dir))
