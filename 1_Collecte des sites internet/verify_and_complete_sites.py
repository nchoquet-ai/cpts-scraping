"""
Vérification et complétion des sites internet CPTS
====================================================
Usage :
  pip install anthropic httpx pandas openpyxl tqdm
  set ANTHROPIC_API_KEY=sk-ant-...        (Windows)
  export ANTHROPIC_API_KEY=sk-ant-...     (Mac/Linux)

  # Lancement complet
  python verify_and_complete_sites.py --input "Liste des CPTS avec site.xlsx"

  # Test sur N lignes
  python verify_and_complete_sites.py --input fichier.xlsx --limit 50

  # Reprendre après interruption
  python verify_and_complete_sites.py --input fichier.xlsx --resume

  # Phases séparées
  python verify_and_complete_sites.py --input fichier.xlsx --verify-only
  python verify_and_complete_sites.py --input fichier.xlsx --search-only

Durée estimée sur 852 CPTS :
  Phase 1 (HTTP)           ~3 min
  Phase 2 (correspondance) ~35 min
  Pause rate limit          ~2 min
  Phase 3 (recherche)      ~30 min (~250 CPTS × 7s)
  Total                    ~1h10
"""

import argparse
import asyncio
import json
import logging
import os
import re
import sys
import time
from collections import Counter
from pathlib import Path
from urllib.parse import urlparse

# Force UTF-8 output on Windows terminals (cp1252 can't encode emoji/box chars)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import anthropic
import httpx
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from tqdm import tqdm

# ─── Configuration ────────────────────────────────────────────────────────────────

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
OUTPUT_DIR        = Path("Output")
INPUT_DIR         = Path("Input")
CLAUDE_MODEL      = "claude-sonnet-4-6"

# HTTP
CONCURRENCY_HTTP  = 20    # vérifications HTTP en parallèle
REQUEST_TIMEOUT   = 12    # secondes par requête
HTTP_RETRIES      = 2     # tentatives sur erreur réseau

# Claude phase 2 (correspondance) — 3 en parallèle respecte le rate limit
CONCURRENCY_MATCH = 3

# Claude phase 3 (recherche) — séquentiel avec délai
SEARCH_DELAY      = 7     # secondes entre chaque appel
SEARCH_RETRIES    = 5     # tentatives sur 429
SEARCH_RETRY_WAIT = 60    # secondes d'attente sur 429

# Pause entre phase 2 et phase 3 pour reset du quota
INTER_PHASE_PAUSE = 120

# Checkpoint phase 3
SAVE_EVERY        = 25

# ─── Logging ──────────────────────────────────────────────────────────────────────
# Note : le dossier output est créé dans run() avant que le FileHandler ne s'ouvre.

def setup_logging():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(OUTPUT_DIR / "run.log", mode="a", encoding="utf-8"),
        ],
    )

log = logging.getLogger(__name__)

# ─── Couleurs Excel ────────────────────────────────────────────────────────────────

COLOR_OK          = "C6EFCE"  # vert        — confirmé
COLOR_OK_UNVERIF  = "E2EFDA"  # vert pâle   — valide, non vérifié
COLOR_PROTECTED   = "FCE4D6"  # orange       — 403
COLOR_UNCERTAIN   = "F4CCCC"  # rose         — incertain
COLOR_WRONG       = "FF7070"  # rouge vif    — mauvaise CPTS
COLOR_DEAD        = "FFC7CE"  # rouge pâle   — mort
COLOR_FOUND       = "D9E1F2"  # bleu pâle    — trouvé, à vérifier
COLOR_NOT_FOUND   = "F2F2F2"  # gris         — introuvable

COLOR_MAP = {
    ("ok",       "oui"):       COLOR_OK,
    ("ok",       "non"):       COLOR_WRONG,
    ("ok",       "incertain"): COLOR_UNCERTAIN,
    ("ok",       "protégé"):   COLOR_PROTECTED,
    ("ok",       "à vérifier"):COLOR_FOUND,
    ("redirect", "oui"):       COLOR_OK,
    ("redirect", "non"):       COLOR_WRONG,
    ("redirect", "incertain"): COLOR_UNCERTAIN,
    ("redirect", "protégé"):   COLOR_PROTECTED,
    ("redirect", "à vérifier"):COLOR_FOUND,
}

# ─── Normalisation URL ─────────────────────────────────────────────────────────────

_INVALID_VALUES = {"", "nan", "none", "n/a", "#n/a", "-", "nd", "nr", "nc"}

def normalize_url(raw: str) -> str:
    """Normalise une URL brute vers https://..."""
    s = str(raw).strip().rstrip("/")
    if s.lower() in _INVALID_VALUES:
        return ""
    if s.startswith(("http://", "https://")):
        return s
    if "." in s:
        return "https://" + s
    return ""

# ─── Parsing JSON robuste (partagé phases 2 et 3) ─────────────────────────────────

def extract_json(text: str):
    """
    Extrait un JSON valide depuis une réponse Claude potentiellement mal formatée.
    4 niveaux de robustesse croissante.
    """
    if not text:
        return None

    # Niveau 1 : parsing direct
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Niveau 2 : nettoyer les backticks markdown
    clean = text.strip()
    if clean.startswith("```"):
        clean = clean.split("\n", 1)[-1] if "\n" in clean else clean[3:]
    if clean.endswith("```"):
        clean = clean.rsplit("```", 1)[0]
    try:
        return json.loads(clean.strip())
    except json.JSONDecodeError:
        pass

    # Niveau 3 : extraire le premier bloc {...} du texte
    brace_start = text.find("{")
    brace_end   = text.rfind("}")
    if brace_start != -1 and brace_end > brace_start:
        try:
            return json.loads(text[brace_start:brace_end + 1])
        except json.JSONDecodeError:
            pass

    # Niveau 4 : extraire une URL https directement (texte libre)
    m = re.search(r"https?://[A-Za-z0-9._~:/?#@!$&'()*+,;=%-]+", text)
    if m:
        url = m.group().rstrip(".,);")
        log.debug(f"URL extraite du texte libre : {url}")
        return {"site_web": url, "confiance": "faible", "source": "URL extraite du texte"}

    return None

# ─── Phase 1 : Vérification HTTP ──────────────────────────────────────────────────

async def _try_url(client: httpx.AsyncClient, url: str) -> dict:
    """Tente de joindre une URL avec retry sur erreurs réseau."""
    for attempt in range(HTTP_RETRIES + 1):
        try:
            r     = await client.get(url, follow_redirects=True, timeout=REQUEST_TIMEOUT)
            final = str(r.url)
            code  = r.status_code

            if code in (401, 403):
                return {"status": "protégé", "final_url": final, "code": code}
            if code in (404, 410):
                return {"status": "dead",    "final_url": final, "code": code}
            if code >= 500:
                if attempt < HTTP_RETRIES:
                    await asyncio.sleep(2 * (attempt + 1))
                    continue
                return {"status": "dead", "final_url": final, "code": code}

            redirected = final.rstrip("/") != url.rstrip("/")
            return {"status": "redirect" if redirected else "ok", "final_url": final, "code": code}

        except (httpx.ConnectTimeout, httpx.ReadTimeout):
            if attempt < HTTP_RETRIES:
                await asyncio.sleep(2 * (attempt + 1))
                continue
            return {"status": "dead", "final_url": url, "code": None, "error": "timeout"}
        except Exception as e:
            return {"status": "dead", "final_url": url, "code": None, "error": str(e)[:80]}

    return {"status": "dead", "final_url": url, "code": None}


async def check_url(client: httpx.AsyncClient, url: str) -> dict:
    """Vérifie une URL. Si 403 sans www, tente la version www."""
    if not url:
        return {"status": "vide", "final_url": "", "code": None}

    result = await _try_url(client, url)

    if result["status"] == "protégé":
        parsed = urlparse(url)
        if not parsed.netloc.startswith("www."):
            www_url    = url.replace(f"://{parsed.netloc}", f"://www.{parsed.netloc}", 1)
            www_result = await _try_url(client, www_url)
            if www_result["status"] in ("ok", "redirect"):
                log.info(f"Version www OK : {www_url}")
                return www_result

    return result


async def verify_all_urls(df: pd.DataFrame) -> pd.DataFrame:
    """Vérifie toutes les URLs en parallèle (phase 1)."""
    urls = df["Site internet"].tolist()
    sem  = asyncio.Semaphore(CONCURRENCY_HTTP)

    async with httpx.AsyncClient(
        headers={
            "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
        },
        follow_redirects=True,
    ) as client:
        async def bounded(url):
            async with sem:
                return await check_url(client, url)

        log.info(f"Phase 1 — Vérification HTTP de {len(urls)} URLs…")
        results = await asyncio.gather(*[bounded(u) for u in urls])

    df = df.copy()
    df["url_status"] = [r["status"]    for r in results]
    df["url_finale"] = [r["final_url"] for r in results]
    df["http_code"]  = [r.get("code")  for r in results]

    counts = Counter(r["status"] for r in results)
    log.info("Phase 1 terminée — " + "  ".join(f"{k}:{v}" for k, v in counts.items()))
    return df

# ─── Phase 2 : Correspondance CPTS ────────────────────────────────────────────────

MATCH_PROMPT = """\
Tu es un expert en santé publique française.
Vérifie si ce site web correspond à la CPTS indiquée.

CPTS attendue : {nom}
URL : {url}

Contenu de la page :
---
{contenu}
---

Réponds UNIQUEMENT avec ce JSON (sans markdown) :
{{"correspondance": "oui"|"non"|"incertain", "raison": "max 100 caractères"}}

- "oui"       : le site mentionne clairement cette CPTS ou un nom très proche
- "non"       : autre organisation, autre CPTS, hors sujet
- "incertain" : site santé de la zone mais identité CPTS non confirmée
"""


async def fetch_page_text(client: httpx.AsyncClient, url: str) -> str:
    """Récupère le texte visible d'une page web."""
    try:
        r    = await client.get(url, timeout=REQUEST_TIMEOUT)
        text = r.text
        text = re.sub(r"<script[^>]*>.*?</script>", " ", text, flags=re.DOTALL)
        text = re.sub(r"<style[^>]*>.*?</style>",   " ", text, flags=re.DOTALL)
        text = re.sub(r"<[^>]+>", " ", text)
        return re.sub(r"\s+", " ", text).strip()[:4000]
    except Exception:
        return ""


def _call_match(claude_client: anthropic.Anthropic, nom: str, url: str, contenu: str) -> dict:
    """Appel Claude synchrone pour vérifier la correspondance (avec parsing robuste + retry 429)."""
    if not contenu:
        return {"correspondance": "incertain", "raison": "contenu inaccessible"}
    for attempt in range(SEARCH_RETRIES + 1):
        try:
            resp = claude_client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=150,
                messages=[{"role": "user", "content":
                            MATCH_PROMPT.format(nom=nom, url=url, contenu=contenu)}],
            )
            # Parsing robuste partagé
            text_blocks = [
                b.text.strip() for b in resp.content
                if hasattr(b, "text") and b.text and b.text.strip()
            ]
            for block in reversed(text_blocks):
                result = extract_json(block)
                if result and "correspondance" in result:
                    return result
            return {"correspondance": "incertain", "raison": "réponse non parseable"}
        except Exception as e:
            if "429" in str(e) and attempt < SEARCH_RETRIES:
                log.warning(
                    f"429 phase 2 — {nom} (tentative {attempt + 1}/{SEARCH_RETRIES})"
                    f" — attente {SEARCH_RETRY_WAIT}s"
                )
                time.sleep(SEARCH_RETRY_WAIT)
                continue
            return {"correspondance": "incertain", "raison": str(e)[:80]}
    return {"correspondance": "incertain", "raison": "échec après retries"}


async def verify_cpts_matches(df: pd.DataFrame, checkpoint_path: Path | None = None) -> pd.DataFrame:
    """Vérifie que chaque URL correspond bien à la bonne CPTS (phase 2)."""
    df = df.copy()
    df["correspondance"]        = ""
    df["correspondance_raison"] = ""

    if not ANTHROPIC_API_KEY:
        log.warning("Clé API absente — correspondance ignorée.")
        return df

    to_check      = df[df["url_status"].isin(["ok", "redirect", "protégé"])].copy()
    claude_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    sem_claude    = asyncio.Semaphore(CONCURRENCY_MATCH)

    log.info(f"Phase 2 — Correspondance sur {len(to_check)} URLs…")

    async with httpx.AsyncClient(
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
        follow_redirects=True,
    ) as http_client:

        async def check_one(idx, row):
            if row["url_status"] == "protégé":
                return idx, {"correspondance": "protégé",
                             "raison": "site bloque les robots — vérifier manuellement"}
            url     = row.get("url_finale") or row.get("Site internet", "")
            contenu = await fetch_page_text(http_client, url)
            async with sem_claude:
                result = await asyncio.get_running_loop().run_in_executor(
                    None, _call_match, claude_client, row["Label_CPTS_OPEX"], url, contenu
                )
            return idx, result

        tasks = [check_one(idx, row) for idx, row in to_check.iterrows()]
        done  = 0
        for coro in tqdm(asyncio.as_completed(tasks), total=len(tasks),
                         desc="Correspondance CPTS"):
            idx, result = await coro
            df.at[idx, "correspondance"]        = result.get("correspondance", "incertain")
            df.at[idx, "correspondance_raison"] = result.get("raison", "")
            done += 1
            # Checkpoint intermédiaire toutes les SAVE_EVERY CPTS
            if checkpoint_path and done % SAVE_EVERY == 0:
                df.to_csv(checkpoint_path, index=False)
                log.info(f"Checkpoint phase 2 : {done}/{len(tasks)}")

    c = df["correspondance"]
    log.info(
        f"Phase 2 terminée — ✅ oui:{(c=='oui').sum()}  ❌ non:{(c=='non').sum()}  "
        f"❓ incertain:{(c=='incertain').sum()}  🔒 protégé:{(c=='protégé').sum()}"
    )
    return df

# ─── Phase 3 : Recherche des sites manquants ──────────────────────────────────────

SEARCH_PROMPT = """\
Tu es un expert en santé publique française.
Trouve le site internet OFFICIEL de cette CPTS :

Nom : {nom}

Effectue une recherche web avec plusieurs requêtes :
- site officiel "{nom}" CPTS
- "{nom}" CPTS France site internet

Réponds UNIQUEMENT avec ce JSON (sans markdown) :
{{
  "site_web":  "https://...",
  "confiance": "haute|moyenne|faible",
  "source":    "explication courte (max 100 caractères)"
}}

- site_web = null si vraiment introuvable
- Retourne UNIQUEMENT le site officiel (pas un annuaire, pas un article)
"""


def _call_search(claude_client: anthropic.Anthropic, nom: str) -> dict:
    """
    Recherche web via Claude.
    Parsing robuste + retry automatique sur 429.
    """
    for attempt in range(SEARCH_RETRIES + 1):
        try:
            resp = claude_client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=500,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=[{"role": "user", "content": SEARCH_PROMPT.format(nom=nom)}],
            )

            # Collecter tous les blocs texte
            text_blocks = [
                b.text.strip() for b in resp.content
                if hasattr(b, "text") and b.text and b.text.strip()
            ]

            # Essayer chaque bloc (le dernier est généralement le bon)
            for block in reversed(text_blocks):
                result = extract_json(block)
                if result and result.get("site_web"):
                    return result

            # Fallback : tenter sur la concaténation de tous les blocs
            combined = " ".join(text_blocks)
            result   = extract_json(combined)
            if result:
                return result

            log.warning(f"Réponse non parseable pour {nom}")
            return {"site_web": None, "confiance": "faible", "source": "réponse non parseable"}

        except Exception as e:
            if "429" in str(e) and attempt < SEARCH_RETRIES:
                log.warning(
                    f"429 — {nom} (tentative {attempt + 1}/{SEARCH_RETRIES})"
                    f" — attente {SEARCH_RETRY_WAIT}s"
                )
                time.sleep(SEARCH_RETRY_WAIT)
                continue
            log.error(f"Recherche échouée {nom}: {e}")
            return {"site_web": None, "confiance": "faible", "source": str(e)[:80]}

    return {"site_web": None, "confiance": "faible", "source": "échec après retries"}


async def search_all_missing(df: pd.DataFrame, checkpoint_path: Path) -> pd.DataFrame:
    """
    Recherche séquentielle des sites manquants / morts / erronés / non trouvés.
    Les sites protégés (403) sont exclus — leur URL existe déjà.
    """
    if not ANTHROPIC_API_KEY:
        log.error("Clé API absente — recherche impossible.")
        return df

    no_url     = df["Site internet"] == ""
    dead       = df["url_status"].isin(["dead", "invalid"])
    wrong      = df["correspondance"] == "non"
    not_found  = df["url_status"] == "non trouvé"   # échecs des runs précédents
    todo       = df[no_url | dead | wrong | not_found].copy()

    log.info(
        f"Phase 3 — {len(todo)} CPTS à rechercher : "
        f"vides:{no_url.sum()}  mortes:{dead.sum()}  "
        f"erronées:{wrong.sum()}  non trouvées:{not_found.sum()}"
    )
    if todo.empty:
        log.info("Phase 3 — Rien à rechercher.")
        return df

    claude_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    df            = df.copy()
    done          = 0

    for idx, row in tqdm(list(todo.iterrows()), total=len(todo), desc="Recherche sites"):

        await asyncio.sleep(SEARCH_DELAY)

        result    = await asyncio.get_running_loop().run_in_executor(
            None, _call_search, claude_client, row["Label_CPTS_OPEX"]
        )
        site      = result.get("site_web") or ""
        confiance = result.get("confiance", "faible")

        if site:
            old_url = row.get("Site internet", "")
            if old_url:
                df.at[idx, "url_originale_morte"] = old_url
            df.at[idx, "Site internet"]         = site
            df.at[idx, "url_status"]            = f"trouvé ({confiance})"
            df.at[idx, "url_finale"]            = site
            df.at[idx, "source_recherche"]      = result.get("source", "")
            df.at[idx, "correspondance"]        = "à vérifier"
            df.at[idx, "correspondance_raison"] = "nouveau site — vérifier manuellement"
            log.info(f"✓ {row['Label_CPTS_OPEX']} → {site} ({confiance})")
        else:
            df.at[idx, "url_status"]       = "non trouvé"
            df.at[idx, "source_recherche"] = result.get("source", "introuvable")
            log.info(f"✗ {row['Label_CPTS_OPEX']} — introuvable")

        done += 1
        # Checkpoint régulier ET à la fin
        if done % SAVE_EVERY == 0 or done == len(todo):
            df.to_csv(checkpoint_path, index=False)
            log.info(f"Checkpoint {done}/{len(todo)}")

    log.info("Phase 3 terminée.")
    return df

# ─── Export Excel ──────────────────────────────────────────────────────────────────

COLS_IN  = [
    "Code_CPTS", "Label_CPTS_OPEX",
    "site_source", "Site internet",
    "url_status", "url_finale", "http_code",
    "correspondance", "correspondance_raison", "source_recherche",
]
COLS_OUT = [
    "Code CPTS", "Nom CPTS",
    "Site source (fichier original)", "Site internet (vérifié/trouvé)",
    "Statut HTTP", "URL finale", "Code HTTP",
    "Correspondance CPTS", "Raison", "Source recherche",
]
COL_WIDTHS = [14, 42, 42, 42, 16, 48, 10, 18, 48, 50]


def export_excel(df: pd.DataFrame, output_path: Path):
    for col in COLS_IN:
        if col not in df.columns:
            df[col] = ""

    df[COLS_IN].rename(columns=dict(zip(COLS_IN, COLS_OUT))).to_excel(
        output_path, index=False
    )
    wb = load_workbook(output_path)
    ws = wb.active

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")

    for letter, width in zip("ABCDEFGHIJ", COL_WIDTHS):
        ws.column_dimensions[letter].width = width

    statut_col = COLS_IN.index("url_status")
    corr_col   = COLS_IN.index("correspondance")

    for row in ws.iter_rows(min_row=2):
        statut = str(row[statut_col].value or "").lower()
        corr   = str(row[corr_col].value   or "").lower()

        if statut in ("dead", "invalid"):
            color = COLOR_DEAD
        elif statut in ("vide", "non trouvé"):
            color = COLOR_NOT_FOUND
        elif statut.startswith("trouvé"):
            color = COLOR_FOUND
        elif statut == "protégé":
            color = COLOR_PROTECTED
        elif statut in ("ok", "redirect"):
            color = COLOR_MAP.get((statut, corr), COLOR_OK_UNVERIF)
        else:
            color = COLOR_NOT_FOUND

        fill = PatternFill("solid", fgColor=color)
        for cell in row:
            cell.fill = fill

    wb.save(output_path)
    log.info(f"Excel exporté : {output_path.resolve()}")

# ─── Statistiques ──────────────────────────────────────────────────────────────────

def print_stats(df: pd.DataFrame):
    s = df.get("url_status",     pd.Series(dtype=str)).fillna("")
    c = df.get("correspondance", pd.Series(dtype=str)).fillna("")
    print("\n─── Résultats ──────────────────────────────────────────")
    print(f"  Total CPTS                 : {len(df)}")
    print(f"  ✅ URLs valides (ok)        : {(s=='ok').sum()}")
    print(f"  🔀 URLs redirigées          : {(s=='redirect').sum()}")
    print(f"  🔒 URLs protégées (403)     : {(s=='protégé').sum()}")
    print(f"  ❌ URLs mortes              : {(s=='dead').sum()}")
    print(f"  🔍 Sites trouvés            : {s.str.startswith('trouvé').sum()}")
    print(f"  ⬜ Sites introuvables       : {(s=='non trouvé').sum()}")
    print(f"  ─── Correspondance ──────────────────────────────────")
    print(f"  ✅ Confirmés                : {(c=='oui').sum()}")
    print(f"  ❌ Mauvaise CPTS (corrigés) : {(c=='non').sum()}")
    print(f"  ❓ Incertains               : {(c=='incertain').sum()}")
    print(f"  🔒 Protégés (manuels)       : {(c=='protégé').sum()}")
    print(f"  🔵 Nouveaux (à vérifier)    : {(c=='à vérifier').sum()}")
    print("────────────────────────────────────────────────────────\n")

# ─── Pipeline principal ────────────────────────────────────────────────────────────

async def run(
    input_path:  str,
    verify_only: bool = False,
    search_only: bool = False,
    limit:       int  = None,
    resume:      bool = False,
):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    checkpoint = OUTPUT_DIR / "checkpoint.csv"
    out_xlsx   = OUTPUT_DIR / "CPTS_sites_verified.xlsx"
    out_csv    = OUTPUT_DIR / "CPTS_sites_verified.csv"

    # ── Chargement ────────────────────────────────────────────────────────────────
    if resume and checkpoint.exists():
        df = pd.read_csv(checkpoint, dtype=str).fillna("")
        log.info(f"Reprise depuis checkpoint ({len(df)} lignes)")
    else:
        if resume:
            log.warning("--resume demandé mais pas de checkpoint — chargement du fichier source")
        df = pd.read_excel(input_path, dtype=str)
        df.columns = [c.strip() for c in df.columns]
        df = df.fillna("")
        if limit:
            df = df.head(limit)
            log.info(f"Mode test : {limit} premières CPTS")
        log.info(f"{len(df)} CPTS chargées depuis {input_path}")

    # ── Initialisation des colonnes ───────────────────────────────────────────────
    for col in ["Code_CPTS", "url_status", "url_finale", "http_code",
                "source_recherche", "url_originale_morte",
                "correspondance", "correspondance_raison"]:
        if col not in df.columns:
            df[col] = ""

    # ── URL source immuable ───────────────────────────────────────────────────────
    if "site_source" not in df.columns:
        df["site_source"] = df["Site internet"].copy()

    # ── Normalisation (toujours, même en --resume ou --search-only) ───────────────
    df["Site internet"] = df["Site internet"].apply(normalize_url)

    # ── Phase 1 + 2 ──────────────────────────────────────────────────────────────
    if not search_only:
        has_url    = df["Site internet"] != ""
        df_with    = df[has_url].copy()
        df_without = df[~has_url].copy()
        df_with    = await verify_all_urls(df_with)
        df         = pd.concat([df_with, df_without]).sort_index()
        df         = await verify_cpts_matches(df, checkpoint)
        df.to_csv(checkpoint, index=False)
        log.info("Checkpoint sauvegardé après phases 1+2")

    # ── Pause inter-phases ────────────────────────────────────────────────────────
    if not verify_only and not search_only:
        log.info(f"Pause {INTER_PHASE_PAUSE}s (reset rate limit)…")
        await asyncio.sleep(INTER_PHASE_PAUSE)

    # ── Phase 3 ───────────────────────────────────────────────────────────────────
    if not verify_only:
        df = await search_all_missing(df, checkpoint)

    # ── Export ────────────────────────────────────────────────────────────────────
    export_excel(df, out_xlsx)
    df.to_csv(out_csv, index=False)
    print_stats(df)
    log.info(f"Fichiers :\n  {out_xlsx.resolve()}\n  {out_csv.resolve()}")

# ─── CLI ───────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Vérification et complétion des sites internet CPTS",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples :
  python verify_and_complete_sites.py --input "Liste des CPTS avec site.xlsx"
  python verify_and_complete_sites.py --input fichier.xlsx --limit 50
  python verify_and_complete_sites.py --input fichier.xlsx --resume
  python verify_and_complete_sites.py --input fichier.xlsx --verify-only
  python verify_and_complete_sites.py --input fichier.xlsx --search-only
        """
    )
    parser.add_argument("--input",       required=True,  help="Fichier Excel source (.xlsx)")
    parser.add_argument("--limit",       type=int,       help="Limiter aux N premières lignes")
    parser.add_argument("--verify-only", action="store_true", help="Phases 1 & 2 uniquement")
    parser.add_argument("--search-only", action="store_true", help="Phase 3 uniquement")
    parser.add_argument("--resume",      action="store_true", help="Reprendre depuis le checkpoint")
    args = parser.parse_args()

    if args.verify_only and args.search_only:
        parser.error("--verify-only et --search-only sont incompatibles.")

    setup_logging()

    asyncio.run(run(
        input_path  = args.input,
        verify_only = args.verify_only,
        search_only = args.search_only,
        limit       = args.limit,
        resume      = args.resume,
    ))
